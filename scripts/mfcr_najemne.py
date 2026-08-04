# Extrakce nájemného z cenové mapy MFČR do data/najemne_mfcr.csv.
#
# ZMĚNA 2026-08-04 (na návrh uživatele, schváleno): dřív se nájem stahoval jen
# pro Prahu scrapingem textových popisků z Leaflet mapy (widget "Mapa_Praha",
# regex na HTML popup labely) — fragilní a navíc bez pokrytí pro Střední Čechy
# (samostatná mapa "Mapa_Praha" logicky nemá středočeská data). Uživatel si
# všiml, že MFČR na stejné stránce nabízí i strukturovaný XLSX "Cenová mapa -
# tabulkové výstupy" pokrývající VŠECHNY kraje ČR — mnohem robustnější zdroj
# (žádné parsování popisků z mapy, čisté sloupce).
#
# Zdroj: https://mf.gov.cz/cs/rozpoctova-politika/podpora-projektoveho-rizeni/
# cenova-mapa/cenova-mapa-infografika — příloha "Cenová mapa - tabulkové
# výstupy" (XLSX). Aktualizuje se 4x ročně, název souboru obsahuje datum
# (např. 2026-05-15_Cenova-mapa.xlsx) — URL se proto hledá dynamicky na
# stránce, nikde není natvrdo.
#
# List "Cenové mapy nájemného": sloupce Kraj/Katastrální území/Obec/Kód obce +
# 4 opakující se bloky po 10 sloupcích (VK1..VK4 = velikostní kategorie bytu:
# 1+kk/1+1, 2+kk/2+1, 3+kk/3+1, 4 a víc pokojů — legenda na listu "Základní
# informace"). Používáme sloupec "Nájemné referenčního bytu za m² v Kč za 1
# měsíc" z každého bloku.
#
# Praha: matchuje se podle Katastrální území (stejná úroveň jako dřív — čtvrť,
# 1:1, žádná agregace).
#
# Střední Čechy: náš model matchuje nájem na úroveň celé OBCE (`ctvrt` u
# středočeských nabídek = název obce, viz src/sreality.py), ale velká města
# mají v XLSX víc katastrálních území (např. Kladno: Kladno, Dubí u Kladna,
# Kročehlavy, Motyčín, Rozdělov, Vrapice, Hnidousy). Agreguje se PROSTÝM
# PRŮMĚREM přes všechna katastrální území patřící té obci — schváleno
# uživatelem 2026-08-04: "pro prvotní screening asi můžeme pracovat s
# průměry a pak při tom detailnějším hledání to nějak zohlední a upraví si
# podle toho koeficient lokality" (tj. přesnější rozlišení čtvrtí uvnitř obce
# je vědomě mimo scope automatického modelu, řeší si uživatel sám ručně).
#
# Kolize stejnojmenných obcí (Roztoky, Jesenice) — STEJNÝ bug jako u cenové
# mapy bytů (viz scripts/stredocesky_cenova_mapa.py, PREDAVACI.md 2026-08-04):
# XLSX má "Roztoky" i "Jesenice" víckrát pod různým "Kód obce" (různé okresy).
# Rozlišeno přes Kód obce (ČSÚ/RUIAN, ověřeno webovým vyhledáváním 2026-08-04):
#   Roztoky (Praha-západ, na našem seznamu velkých měst) = kód obce 539627
#   Jesenice (Praha-západ, na našem seznamu velkých měst) = kód obce 539325
# Jakákoli JINÁ, dosud neznámá kolize (mimo tyhle dvě) se ohlásí a zastaví
# skript — žádná tichá volba "prvního nalezeného".
import csv
import io
import re

import openpyxl
import requests

UA = {"User-Agent": "Mozilla/5.0"}
BASE = "https://mf.gov.cz"
INFO_URL = BASE + "/cs/rozpoctova-politika/podpora-projektoveho-rizeni/cenova-mapa/cenova-mapa-infografika"
MESTA_STC_CSV = "data/mesta_stredocechy.csv"
OUT_CSV = "data/najemne_mfcr.csv"
LIST_NAJEMNE = "Cenové mapy nájemného"

# Stejné popisky sloupců, jaké už čeká src/valuation.py (_SKUPINY) — beze změny,
# ať se nemusí nic upravovat na straně spotřebitele tohohle CSV.
SKUPINY = ["1+kk, 1+1", "2+kk, 2+1", "3+kk, 3+1", "4+kk, 4+1"]

KOD_OBCE_SPRAVNY = {"roztoky": 539627, "jesenice": 539325}


def _norm(s):
    return s.strip().lower().replace(" ", "_").replace("-", "_")


def nacti_velka_mesta() -> set:
    with open(MESTA_STC_CSV, encoding="utf-8-sig") as f:
        return {_norm(r["nazev"]) for r in csv.DictReader(f)}


def stahni_xlsx_url() -> str:
    r = requests.get(INFO_URL, headers=UA, timeout=60)
    r.raise_for_status()
    m = re.search(r'(?:https://mf\.gov\.cz)?(/assets/attachments/[^\s")]*Cenova-mapa\.xlsx)', r.text)
    if not m:
        raise RuntimeError("Odkaz na XLSX cenové mapy nájemného nenalezen na stránce MFČR "
                            "— zastavuji, nenahrazuji tichým odhadem (možná změna struktury stránky).")
    return BASE + m.group(1)


def stahni_radky(url: str) -> list:
    r = requests.get(url, headers=UA, timeout=120)
    r.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)
    if LIST_NAJEMNE not in wb.sheetnames:
        raise RuntimeError(f"List '{LIST_NAJEMNE}' nenalezen v XLSX (listy: {wb.sheetnames}) "
                            f"— MFČR pravděpodobně změnilo strukturu, zastavuji.")
    ws = wb[LIST_NAJEMNE]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    if header[0] != "Kraj" or header[1] != "Katastrální území" or header[2] != "Obec" or header[3] != "Kód obce":
        raise RuntimeError(f"Neočekávaná hlavička listu '{LIST_NAJEMNE}' ({header[:4]}) "
                            f"— MFČR pravděpodobně změnilo strukturu, zastavuji.")
    return list(it)


def _hodnoty_vk(radek) -> list:
    """[VK1, VK2, VK3, VK4] hodnoty sloupce 'Nájemné referenčního bytu za m² v
    Kč za 1 měsíc' — bloky po 10 sloupcích (VK + 8 statistik + 1 mezera) od
    indexu 4; sloupec s nájemným je 2. v bloku (index bloku + 1)."""
    vysledky = []
    for blok in range(4):
        idx = 4 + blok * 10 + 1
        v = radek[idx] if idx < len(radek) else None
        vysledky.append(float(v) if isinstance(v, (int, float)) else None)
    return vysledky


def zpracuj():
    velka_mesta = nacti_velka_mesta()
    url = stahni_xlsx_url()
    print("XLSX cenové mapy nájemného:", url)
    radky = stahni_radky(url)
    print(f"Načteno {len(radky)} řádků (celá ČR).")

    praha = {}            # norm(katastrální území) -> [VK1..VK4]
    stc_soucty = {}       # norm(obec) -> [[suma, počet], ...] pro VK1..VK4
    stc_null_ku_pocet = {}  # norm(obec) -> počet řádků BEZ katastrálního území

    # POZOR: víc řádků se stejnou "Obec" a RŮZNÝM "Kód obce" je běžné a v pořádku
    # u měst rozdělených na víc katastrálních území (např. Kladno má 7 k.ú., každé
    # se svým kódem k.ú. — to NENÍ kolize, jen se to sčítá do průměru za obec).
    # Skutečná kolize stejnojmenných-ale-jiných obcí (Roztoky, Jesenice) se pozná
    # jinak: obě sdílené obce v tomhle XLSX mají "Katastrální území" = null (jsou
    # to malé, dál nedělené obce) — víc než 1 řádek se stejným jménem obce A
    # PRÁZDNÝM katastrálním územím je signatura kolize, ověřeno 2026-08-04 na
    # celém Středočeském kraji (49 takových jmen, žádné jiné z našich 43 měst).
    for r in radky:
        kraj, kat_uzemi, obec, kod_obce = r[0], r[1], r[2], r[3]
        vk = _hodnoty_vk(r)
        if kraj == "Hlavní město Praha" and kat_uzemi:
            praha[_norm(kat_uzemi)] = vk
        elif kraj == "Středočeský kraj" and obec:
            klic = _norm(obec)
            if klic not in velka_mesta:
                continue  # menší obce mimo náš seznam velkých měst (>5000 obyv.) ignorujeme
            spravny_kod = KOD_OBCE_SPRAVNY.get(klic)
            if spravny_kod and kod_obce != spravny_kod:
                continue  # jiná stejnojmenná obec v jiném okrese (viz komentář výše) — přeskočit
            if kat_uzemi is None:
                stc_null_ku_pocet[klic] = stc_null_ku_pocet.get(klic, 0) + 1
            radky_klic = stc_soucty.setdefault(klic, [[0.0, 0] for _ in range(4)])
            for i, v in enumerate(vk):
                if v is not None:
                    radky_klic[i][0] += v
                    radky_klic[i][1] += 1

    nova_kolize = {k: n for k, n in stc_null_ku_pocet.items() if n > 1 and k not in KOD_OBCE_SPRAVNY}
    if nova_kolize:
        raise RuntimeError(f"Nová (dosud neřešená) kolize stejnojmenných obcí ve Středních "
                            f"Čechách: {nova_kolize} — zastavuji, nenahrazuji tichým "
                            f"výběrem, je potřeba ručně ověřit správný Kód obce (ČSÚ/RUIAN) "
                            f"a doplnit do KOD_OBCE_SPRAVNY.")

    stc = {klic: [s / n if n else None for s, n in vk] for klic, vk in stc_soucty.items()}

    chybi = velka_mesta - set(stc)
    if chybi:
        print(f"POZOR: {len(chybi)} velkých měst ze seznamu nemá řádek v mapě nájemného: {sorted(chybi)}")

    vysledek = {**praha, **stc}
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["ctvrt"] + SKUPINY)
        for ctvrt in sorted(vysledek):
            w.writerow([ctvrt] + [v if v is not None else "" for v in vysledek[ctvrt]])

    print(f"Zapsáno {len(vysledek)} lokalit do {OUT_CSV} "
          f"(Praha: {len(praha)} katastrálních území, Střední Čechy: {len(stc)} obcí "
          f"z {len(velka_mesta)} na seznamu).")


if __name__ == "__main__":
    zpracuj()
